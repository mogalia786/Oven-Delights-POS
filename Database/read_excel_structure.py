import openpyxl

# Read the Excel file
wb = openpyxl.load_workbook(r'C:\Development Apps\Cascades projects\Oven-Delights-ERP\Oven-Delights-ERP\Oven-Delights-ERP\Documentation\StockItems with Prices and descriptions\Updated Inventory Pricing_.xlsx')

# Get the first sheet
sheet = wb.active

# Print sheet name
print(f"Sheet name: {sheet.title}")
print("\n" + "="*80)

# Print headers (first row)
print("\nHeaders:")
headers = []
for col in range(1, sheet.max_column + 1):
    cell_value = sheet.cell(row=1, column=col).value
    headers.append(cell_value)
    print(f"Column {col}: {cell_value}")

print("\n" + "="*80)
print("\nFirst 10 rows of data:")
print("-"*80)

# Print first 10 rows
for row in range(1, min(11, sheet.max_row + 1)):
    row_data = []
    for col in range(1, min(10, sheet.max_column + 1)):  # Show first 9 columns
        cell_value = sheet.cell(row=row, column=col).value
        row_data.append(str(cell_value)[:30] if cell_value else "")
    print(f"Row {row}: {' | '.join(row_data)}")

print("\n" + "="*80)
print(f"\nTotal rows: {sheet.max_row}")
print(f"Total columns: {sheet.max_column}")
