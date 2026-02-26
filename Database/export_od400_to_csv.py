import openpyxl
import csv

# Read the Excel file
wb = openpyxl.load_workbook(r'C:\Development Apps\Cascades projects\Oven-Delights-ERP\Oven-Delights-ERP\Oven-Delights-ERP\Documentation\StockItems with Prices and descriptions\Updated Inventory Pricing_.xlsx')

# Get the OD400 - Umhlanga sheet
sheet = wb['OD400 - Umhlanga']

# Output CSV file
output_file = r'c:\Development Apps\Cascades projects\Overn-Delights-POS\Overn-Delights-POS\Database\OD400_Umhlanga_Prices.csv'

print(f"Exporting OD400 - Umhlanga prices to CSV...")
print(f"Total rows: {sheet.max_row}")

# Write to CSV
with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
    writer = csv.writer(csvfile)
    
    # Write header
    headers = ['ItemCode', 'Barcode', 'ItemDescription', 'Category', 'ItemCategory', 'Ingredients', 'Description', 'Whse', 'Cost', 'InclPrice']
    writer.writerow(headers)
    
    # Write data rows (skip header row)
    rows_written = 0
    rows_with_prices = 0
    
    for row in range(2, sheet.max_row + 1):
        item_code = sheet.cell(row=row, column=1).value
        barcode = sheet.cell(row=row, column=2).value
        item_desc = sheet.cell(row=row, column=3).value
        category = sheet.cell(row=row, column=4).value
        item_cat = sheet.cell(row=row, column=5).value
        ingredients = sheet.cell(row=row, column=6).value
        description = sheet.cell(row=row, column=7).value
        whse = sheet.cell(row=row, column=8).value
        cost = sheet.cell(row=row, column=9).value
        incl_price = sheet.cell(row=row, column=10).value
        
        # Only write rows with valid item code and price
        if item_code and incl_price:
            writer.writerow([
                item_code,
                barcode or '',
                item_desc or '',
                category or '',
                item_cat or '',
                ingredients or '',
                description or '',
                whse or '',
                cost or 0,
                incl_price or 0
            ])
            rows_written += 1
            if incl_price and float(incl_price) > 0:
                rows_with_prices += 1

print(f"\nExport complete!")
print(f"Total rows written: {rows_written}")
print(f"Rows with prices: {rows_with_prices}")
print(f"Output file: {output_file}")
