import openpyxl

# Read the Excel file
wb = openpyxl.load_workbook(r'C:\Development Apps\Cascades projects\Oven-Delights-ERP\Oven-Delights-ERP\Oven-Delights-ERP\Documentation\StockItems with Prices and descriptions\Updated Inventory Pricing_.xlsx')

print("All sheets in the workbook:")
print("="*80)
for idx, sheet_name in enumerate(wb.sheetnames, 1):
    sheet = wb[sheet_name]
    print(f"{idx}. {sheet_name} - {sheet.max_row} rows, {sheet.max_column} columns")

print("\n" + "="*80)
print("\nChecking for price data in each sheet...")
print("-"*80)

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    
    # Find Cost and Price columns
    cost_col = None
    price_col = None
    whse_col = None
    
    for col in range(1, sheet.max_column + 1):
        header = str(sheet.cell(row=1, column=col).value).lower() if sheet.cell(row=1, column=col).value else ""
        if 'cost' in header:
            cost_col = col
        if 'price' in header or 'incl' in header:
            price_col = col
        if 'whse' in header or 'warehouse' in header:
            whse_col = col
    
    # Check if there's actual price data
    has_prices = False
    if price_col:
        for row in range(2, min(20, sheet.max_row + 1)):
            price_val = sheet.cell(row=row, column=price_col).value
            if price_val and price_val != 0:
                has_prices = True
                break
    
    whse_code = sheet.cell(row=2, column=whse_col).value if whse_col else "N/A"
    
    print(f"\nSheet: {sheet_name}")
    print(f"  Warehouse Code: {whse_code}")
    print(f"  Cost Column: {cost_col}")
    print(f"  Price Column: {price_col}")
    print(f"  Has Price Data: {has_prices}")
